"""
app/api/v1/endpoints/inventory_turnover.py
──────────────────────────────────────────
REST endpoints untuk Inventory Turnover SCM.
Gunakan DB session publik (bukan plant schema) karena data ada di national_sc.

Daftarkan di app/api/v1/router.py:
    from app.api.v1.endpoints import inventory_turnover
    api_router.include_router(inventory_turnover.router)
"""

import io
import calendar
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.core.deps import CurrentUser
from app.services import inventory_turnover as svc

router = APIRouter(prefix="/inventory-turnover", tags=["Inventory Turnover SCM"])


# ─── helpers ─────────────────────────────────────────────────────────────────

def _parse_csv(raw: Optional[str]) -> list[str]:
    return [v for v in (raw or "").split(",") if v] if raw else []


def _parse_csv_int(raw: Optional[str]) -> list[int]:
    try:
        return [int(v) for v in (raw or "").split(",") if v]
    except ValueError:
        raise HTTPException(400, "Parameter bulan harus angka dipisah koma")


# ─── endpoints ────────────────────────────────────────────────────────────────

@router.get("/filter-options")
def filter_options(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
):
    return svc.get_filter_options(db)


@router.get("/dashboard")
def dashboard(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    year:   int            = Query(...),
    months: Optional[str] = Query(None),
    plants: Optional[str] = Query(None),
):
    return svc.get_dashboard_summary(
        db, year,
        months=_parse_csv_int(months),
        plants=_parse_csv(plants),
    )


@router.get("/trend-analysis")
def trend_analysis(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    year:   int            = Query(...),
    months: Optional[str] = Query(None),
):
    return svc.get_trend_analysis(db, year, months=_parse_csv_int(months))


@router.get("/plant-performance")
def plant_performance(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    year:          int            = Query(...),
    plants:        Optional[str] = Query(None),
    business_unit: Optional[str] = Query(None),
):
    return svc.get_monthly_performance(
        db, year,
        plants=_parse_csv(plants),
        business_unit=_parse_csv(business_unit),
    )


@router.get("/plant-comparison")
def plant_comparison(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    year:   int            = Query(...),
    months: Optional[str] = Query(None),
    plants: Optional[str] = Query(None),
):
    return svc.get_plant_comparison(
        db, year,
        months=_parse_csv_int(months),
        plants=_parse_csv(plants),
    )


@router.get("/submission-status")
def submission_status(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    year:   int            = Query(...),
    months: Optional[str] = Query(None),
    plants: Optional[str] = Query(None),
):
    return svc.get_submission_status(
        db, year,
        months=_parse_csv_int(months),
        plants=_parse_csv(plants),
    )


# ─── Template download ────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill(fill_type="solid", fgColor="6691DC")
_HEADER_FONT   = Font(bold=True, color="000000", size=11)
_BORDER_SIDE   = Side(style="thin", color="E1EBFC")
_THIN_BORDER   = Border(top=_BORDER_SIDE, left=_BORDER_SIDE, bottom=_BORDER_SIDE, right=_BORDER_SIDE)
_ALIGN_LEFT     = Alignment(vertical="center", horizontal="left")

_TEMPLATE_COLS = [
    ("year", 10), ("month", 10), ("plant", 15), ("plant_code", 15),
    ("business_unit", 20), ("mtyp", 15), ("material_code", 25),
    ("material_desc", 50), ("kind_of_product", 30), ("uom", 15),
    ("saldo_awal", 20), ("penerimaan", 20), ("pemakaian", 20), ("saldo_akhir", 20),
]


@router.get("/template")
def download_template(current_user: CurrentUser):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Turnover"
    for col_idx, (header, width) in enumerate(_TEMPLATE_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill, cell.font, cell.border, cell.alignment = (
            _HEADER_FILL, _HEADER_FONT, _THIN_BORDER, _ALIGN_LEFT
        )
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Inventory_Turnover_Template.xlsx"},
    )


# ─── Upload Excel ─────────────────────────────────────────────────────────────

def _clean(val):
    return val


def _safe_num(val) -> float:
    try:
        n = float(val)
        return 0.0 if (n != n or abs(n) == float("inf")) else n
    except (TypeError, ValueError):
        return 0.0


def _sanitize(val) -> str:
    return " ".join(str(val).strip().split()) if val is not None else ""


@router.post("/upload")
def upload_inventory(
    current_user: CurrentUser,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    """Upload file Excel ke national_sc.inventory_turnover."""
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(400, "Hanya file .xlsx yang diperbolehkan.")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "File tidak valid atau korup.")

    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Sheet tidak ditemukan.")

    rows_to_insert = []
    keys_to_delete = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 14:
            raise HTTPException(400, f"Baris ke-{row_num} tidak lengkap (kurang kolom).")
        raw = [row[i] for i in range(14)]
        if any(v is None or v == "" for v in raw):
            raise HTTPException(400, f"Upload dibatalkan. Baris ke-{row_num} tidak lengkap.")

        row_year  = int(raw[0])
        row_month = int(raw[1])
        plant         = _sanitize(raw[2]).upper()
        plant_code    = _sanitize(raw[3]).upper()
        business_unit = _sanitize(raw[4]).lower()
        mtyp          = _sanitize(raw[5]).upper()
        material_code = _sanitize(raw[6]).upper()
        material_desc = _sanitize(raw[7])
        kop           = _sanitize(raw[8])
        uom           = _sanitize(raw[9]).upper()

        saldo_awal  = _safe_num(raw[10])
        penerimaan  = _safe_num(raw[11])
        pemakaian   = _safe_num(raw[12])
        saldo_akhir = _safe_num(raw[13])

        avg_saldo = round((saldo_awal + saldo_akhir) / 2)
        turnover  = 0 if avg_saldo == 0 else round(pemakaian / avg_saldo)
        days_in_m = calendar.monthrange(row_year, row_month)[1]
        doi       = 0 if turnover == 0 else round(days_in_m / turnover)

        rows_to_insert.append((
            row_year, row_month, plant, plant_code, business_unit,
            mtyp, material_code, material_desc, kop, uom,
            saldo_awal, penerimaan, pemakaian, saldo_akhir,
            avg_saldo, turnover, doi,
        ))
        keys_to_delete.append((row_year, row_month, plant, business_unit, material_code))

    if not rows_to_insert:
        raise HTTPException(400, "File kosong atau tidak ada data.")

    from sqlalchemy import text
    try:
        # Delete existing
        db.execute(text("""
            DELETE FROM national_sc.inventory_turnover
            WHERE (year, month, plant, business_unit, material_code)
            IN (SELECT unnest(:years::int[]), unnest(:months::int[]),
                       unnest(:plants::text[]), unnest(:bus::text[]),
                       unnest(:matcodes::text[]))
        """), {
            "years":    [k[0] for k in keys_to_delete],
            "months":   [k[1] for k in keys_to_delete],
            "plants":   [k[2] for k in keys_to_delete],
            "bus":      [k[3] for k in keys_to_delete],
            "matcodes": [k[4] for k in keys_to_delete],
        })

        # Insert
        for r in rows_to_insert:
            db.execute(text("""
                INSERT INTO national_sc.inventory_turnover
                (year, month, plant, plant_code, business_unit, mtyp, material_code,
                 material_desc, kind_of_product, uom, saldo_awal, penerimaan, pemakaian,
                 saldo_akhir, avg_saldo, turnover, days_of_inventory, created_at)
                VALUES (:yr,:mo,:pl,:pc,:bu,:mt,:mc,:md,:kp,:um,
                        :sa,:pn,:pm,:sk,:as_,:tr,:doi,NOW())
            """), dict(
                yr=r[0], mo=r[1], pl=r[2], pc=r[3], bu=r[4],
                mt=r[5], mc=r[6], md=r[7], kp=r[8], um=r[9],
                sa=r[10], pn=r[11], pm=r[12], sk=r[13],
                as_=r[14], tr=r[15], doi=r[16],
            ))
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Gagal menyimpan data: {str(e)}")

    return {"message": f"Berhasil! {len(rows_to_insert)} baris telah diupload."}
