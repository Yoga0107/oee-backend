import io
import uuid
from datetime import datetime as dt

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.database import get_plant_db
from app.core.deps import CurrentUser, CurrentPlant
from app.models.plant_schema import (
    ProductionOutput, OUTPUT_TYPE_CATEGORY, OUTPUT_TYPES,
    MachineLossInput, MachineLossLvl1, MachineLossLvl2, MachineLossLvl3,
    MasterLine, MasterShift, MasterFeedCode, MasterOutputType, DEFAULT_OUTPUT_TYPES,
)
from app.schemas.input import (
    ProductionOutputCreate, ProductionOutputUpdate, ProductionOutputResponse,
    MachineLossInputCreate, MachineLossInputUpdate, MachineLossInputResponse,
)

router = APIRouter(prefix="/input", tags=["Input Data"])

EXPORT_ROLES = {"administrator", "plant_manager", "operator"}
IMPORT_ROLES = {"administrator", "plant_manager"}


def _assert_role(user, allowed_roles: set, action: str = "action"):
    role_name = user.role.name if user.role else ""
    if not user.is_superuser and role_name not in allowed_roles:
        raise HTTPException(
            status_code=403,
            detail=f"Akses ditolak: {action} hanya untuk {', '.join(sorted(allowed_roles))}",
        )


def _db(plant: CurrentPlant) -> Session:
    return next(get_plant_db(plant.schema_name))


def _get_output_types(db) -> list:
    """Return active output types ordered by sort_order; seed defaults if empty."""
    if db.query(MasterOutputType).count() == 0:
        for d in DEFAULT_OUTPUT_TYPES:
            db.add(MasterOutputType(**d))
        db.commit()
    return (
        db.query(MasterOutputType)
        .filter(MasterOutputType.is_active == True)
        .order_by(MasterOutputType.sort_order, MasterOutputType.id)
        .all()
    )


def _build_group_response(rows: list) -> dict:
    """
    Flatten a list of ProductionOutput rows sharing the same group_id into
    a single response dict. quantities berisi {output_type_code: quantity}.
    """
    if not rows:
        return {}
    ref = rows[0]
    quantities = {r.output_type: r.quantity for r in rows}
    actual_output = sum(quantities.values())
    return {
        "id":              ref.group_id,
        "date":            ref.date,
        "line_id":         ref.line_id,
        "shift_id":        ref.shift_id,
        "feed_code_id":    ref.feed_code_id,
        "production_plan": ref.production_plan,
        "quantities":      quantities,
        "actual_output":   actual_output,
        "remarks":         ref.remarks,
        "is_active":       ref.is_active,
        "created_at":      ref.created_at,
        "created_by_id":   ref.created_by_id,
        "line_name":       ref.line.name      if ref.line      else None,
        "shift_name":      ref.shift.name     if ref.shift     else None,
        "feed_code_code":  ref.feed_code.code if ref.feed_code else None,
    }


def _get_group(db, group_id: str) -> list:
    """Return active rows for a group_id, ordered by output_type."""
    return (
        db.query(ProductionOutput)
        .filter(ProductionOutput.group_id == group_id, ProductionOutput.is_active == True)
        .all()
    )


def _build_loss_response(item):
    return {
        "id": item.id, "date": item.date,
        "line_id": item.line_id, "shift_id": item.shift_id,
        "feed_code_id": item.feed_code_id,
        "loss_l1_id": item.loss_l1_id, "loss_l2_id": item.loss_l2_id, "loss_l3_id": item.loss_l3_id,
        "time_from": item.time_from, "time_to": item.time_to,
        "duration_minutes": item.duration_minutes, "remarks": item.remarks,
        "is_active": item.is_active, "created_at": item.created_at,
        "created_by_id": item.created_by_id,
        "line_name": item.line.name if item.line else None,
        "shift_name": item.shift.name if item.shift else None,
        "feed_code_code": item.feed_code.code if item.feed_code else None,
        "loss_l1_name": item.loss_l1.name if item.loss_l1 else None,
        "loss_l2_name": item.loss_l2.name if item.loss_l2 else None,
        "loss_l3_name": item.loss_l3.name if item.loss_l3 else None,
    }



def _validate_loss_refs(db, d):
    if d.get("line_id"):
        if not db.query(MasterLine).filter(MasterLine.id == d["line_id"], MasterLine.is_active == True).first():
            raise HTTPException(404, "Line not found")
    if d.get("shift_id"):
        if not db.query(MasterShift).filter(MasterShift.id == d["shift_id"], MasterShift.is_active == True).first():
            raise HTTPException(404, "Shift not found")
    if d.get("feed_code_id"):
        if not db.query(MasterFeedCode).filter(MasterFeedCode.id == d["feed_code_id"], MasterFeedCode.is_active == True).first():
            raise HTTPException(404, "Feed code not found")
    if d.get("loss_l1_id"):
        if not db.query(MachineLossLvl1).filter(MachineLossLvl1.machine_losses_lvl_1_id == d["loss_l1_id"]).first():
            raise HTTPException(404, "Loss L1 not found")
    if d.get("loss_l2_id"):
        if not db.query(MachineLossLvl2).filter(MachineLossLvl2.machine_losses_lvl_2_id == d["loss_l2_id"]).first():
            raise HTTPException(404, "Loss L2 not found")
    if d.get("loss_l3_id"):
        if not db.query(MachineLossLvl3).filter(MachineLossLvl3.machine_losses_lvl_3_id == d["loss_l3_id"]).first():
            raise HTTPException(404, "Loss L3 not found")


# ─── Excel style helpers ──────────────────────────────────────────────────────

def _border(style="thin"):
    s = Side(style=style, color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg, fc="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fc, size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("medium")


def _data(cell, row_i, numfmt=None, center=False):
    bg = "F8FAFC" if row_i % 2 == 0 else "FFFFFF"
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(size=9, name="Calibri")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = _border()
    if numfmt:
        cell.number_format = numfmt


def _title_rows(ws, title, ncols, bg):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=f"Generated: {dt.now().strftime('%d %B %Y  %H:%M')}")
    c2.fill = PatternFill("solid", fgColor="1E3A5F")
    c2.font = Font(color="BFD7FF", size=9, italic=True, name="Calibri")
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18


# ════════════════════════════════════════════════════════
# PRODUCTION OUTPUTS – CRUD
# ════════════════════════════════════════════════════════

@router.get("/production-outputs", response_model=list[ProductionOutputResponse])
def list_production_outputs(current_user: CurrentUser, plant: CurrentPlant,
    date_from: str | None = None, date_to: str | None = None,
    line_id: int | None = None, shift_id: int | None = None):
    db = _db(plant)
    q = db.query(ProductionOutput).filter(ProductionOutput.is_active == True)
    if date_from: q = q.filter(ProductionOutput.date >= dt.fromisoformat(date_from))
    if date_to:   q = q.filter(ProductionOutput.date <= dt.fromisoformat(date_to + "T23:59:59"))
    if line_id:   q = q.filter(ProductionOutput.line_id == line_id)
    if shift_id:  q = q.filter(ProductionOutput.shift_id == shift_id)
    rows = q.order_by(ProductionOutput.date.desc(), ProductionOutput.group_id).all()

    # Auto-backfill rows with NULL group_id (data lama sebelum migration dijalankan).
    needs_commit = False
    for r in rows:
        if not r.group_id:
            r.group_id = str(uuid.uuid4())
            if not r.output_type:
                r.output_type = "finished_goods"
                r.category    = "FG"
                r.quantity     = 0
            needs_commit = True
    if needs_commit:
        db.commit()
        for r in rows: db.refresh(r)

    # Group rows by group_id preserving date-desc order
    seen: dict[str, list] = {}
    order: list[str] = []
    for r in rows:
        if r.group_id not in seen:
            seen[r.group_id] = []
            order.append(r.group_id)
        seen[r.group_id].append(r)
    return [_build_group_response(seen[gid]) for gid in order]


@router.post("/production-outputs", response_model=ProductionOutputResponse, status_code=201)
def create_production_output(payload: ProductionOutputCreate, current_user: CurrentUser, plant: CurrentPlant):
    db = _db(plant)
    if not db.query(MasterLine).filter(MasterLine.id == payload.line_id, MasterLine.is_active == True).first():
        raise HTTPException(404, "Line tidak ditemukan")
    if not db.query(MasterShift).filter(MasterShift.id == payload.shift_id, MasterShift.is_active == True).first():
        raise HTTPException(404, "Shift tidak ditemukan")
    if payload.feed_code_id:
        if not db.query(MasterFeedCode).filter(MasterFeedCode.id == payload.feed_code_id, MasterFeedCode.is_active == True).first():
            raise HTTPException(404, "Kode pakan tidak ditemukan")
    # 1 submit = N rows (1 per active output type), semua sharing group_id
    output_types = _get_output_types(db)
    if not output_types:
        raise HTTPException(400, "Tidak ada output type aktif. Konfigurasi master output type terlebih dahulu.")
    gid = str(uuid.uuid4())
    new_rows = []
    for ot in output_types:
        row = ProductionOutput(
            group_id=gid,
            date=payload.date, line_id=payload.line_id, shift_id=payload.shift_id,
            feed_code_id=payload.feed_code_id, production_plan=payload.production_plan,
            output_type=ot.code,
            category=ot.category,
            quantity=payload.quantities.get(ot.code, 0),
            remarks=payload.remarks,
            created_by_id=current_user.id,
        )
        db.add(row)
        new_rows.append(row)
    db.commit()
    for r in new_rows: db.refresh(r)
    return _build_group_response(new_rows)


@router.put("/production-outputs/{group_id}", response_model=ProductionOutputResponse)
def update_production_output(group_id: str, payload: ProductionOutputUpdate, current_user: CurrentUser, plant: CurrentPlant):
    """Update semua row dalam satu group (group_id). group_id = id dari response."""
    db = _db(plant)
    rows = _get_group(db, group_id)
    if not rows: raise HTTPException(404, "Production output tidak ditemukan")
    data = payload.model_dump(exclude_none=True)
    if "line_id" in data and not db.query(MasterLine).filter(MasterLine.id == data["line_id"], MasterLine.is_active == True).first():
        raise HTTPException(404, "Line tidak ditemukan")
    if "shift_id" in data and not db.query(MasterShift).filter(MasterShift.id == data["shift_id"], MasterShift.is_active == True).first():
        raise HTTPException(404, "Shift tidak ditemukan")
    if "feed_code_id" in data and data["feed_code_id"]:
        if not db.query(MasterFeedCode).filter(MasterFeedCode.id == data["feed_code_id"], MasterFeedCode.is_active == True).first():
            raise HTTPException(404, "Kode pakan tidak ditemukan")
    # Shared fields — apply to all rows in the group
    shared_fields = {"date", "line_id", "shift_id", "feed_code_id", "production_plan", "remarks"}
    new_quantities = data.pop("quantities", None)  # {output_type_code: quantity}
    for row in rows:
        for k, v in data.items():
            if k in shared_fields:
                setattr(row, k, v)
        # Update quantity dinamis dari quantities dict
        if new_quantities is not None and row.output_type in new_quantities:
            row.quantity = new_quantities[row.output_type]
        row.updated_by_id = current_user.id
    db.commit()
    for r in rows: db.refresh(r)
    return _build_group_response(rows)


@router.delete("/production-outputs/{group_id}", status_code=204)
def delete_production_output(group_id: str, current_user: CurrentUser, plant: CurrentPlant):
    """Soft-delete semua row dalam satu group."""
    db = _db(plant)
    rows = db.query(ProductionOutput).filter(ProductionOutput.group_id == group_id).all()
    if not rows: raise HTTPException(404, "Production output tidak ditemukan")
    for row in rows:
        row.is_active = False
        row.updated_by_id = current_user.id
    db.commit()


@router.get("/production-outputs/by-type", response_model=list[dict])
def list_output_by_type(
    current_user: CurrentUser,
    plant: CurrentPlant,
    date_from:   str | None = None,
    date_to:     str | None = None,
    line_id:     int | None = None,
    shift_id:    int | None = None,
    output_type: str | None = None,
):
    """
    Return production output rows per type — each row is one output_type entry.
    Now reads directly from production_outputs (no join to a separate items table).
    """
    db = _db(plant)
    q = db.query(ProductionOutput).filter(ProductionOutput.is_active == True)
    if date_from:   q = q.filter(ProductionOutput.date >= dt.fromisoformat(date_from))
    if date_to:     q = q.filter(ProductionOutput.date <= dt.fromisoformat(date_to + "T23:59:59"))
    if line_id:     q = q.filter(ProductionOutput.line_id == line_id)
    if shift_id:    q = q.filter(ProductionOutput.shift_id == shift_id)
    if output_type: q = q.filter(ProductionOutput.output_type == output_type)
    rows = q.order_by(ProductionOutput.date.desc(), ProductionOutput.group_id).all()
    return [{
        "item_id":        r.id,
        "group_id":       r.group_id,
        "date":           r.date,
        "line_id":        r.line_id,
        "line_name":      r.line.name  if r.line      else None,
        "shift_id":       r.shift_id,
        "shift_name":     r.shift.name if r.shift     else None,
        "feed_code_id":   r.feed_code_id,
        "feed_code_code": r.feed_code.code if r.feed_code else None,
        "output_type":    r.output_type,
        "category":       r.category,
        "quantity":       r.quantity,
        "remarks":        r.remarks,
    } for r in rows]


# ════════════════════════════════════════════════════════
# PRODUCTION OUTPUTS – EXPORT
# ════════════════════════════════════════════════════════

@router.get("/production-outputs/export")
def export_production_outputs(current_user: CurrentUser, plant: CurrentPlant,
    date_from: str | None = None, date_to: str | None = None,
    line_id: int | None = None, shift_id: int | None = None):
    """Export Production Outputs to Excel. Roles: administrator, plant_manager, operator."""
    _assert_role(current_user, EXPORT_ROLES, "Export Production Output")

    db = _db(plant)
    q = db.query(ProductionOutput).filter(ProductionOutput.is_active == True)
    if date_from: q = q.filter(ProductionOutput.date >= dt.fromisoformat(date_from))
    if date_to:   q = q.filter(ProductionOutput.date <= dt.fromisoformat(date_to + "T23:59:59"))
    if line_id:   q = q.filter(ProductionOutput.line_id == line_id)
    if shift_id:  q = q.filter(ProductionOutput.shift_id == shift_id)
    items = q.order_by(ProductionOutput.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Production Output"
    NCOLS = 15

    _title_rows(ws, f"Production Output  |  {plant.name}", NCOLS, "1D4ED8")

    # KPI strip (row 3)
    # Build grouped data for KPI and export rows
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    grp_order: list[str] = []
    for r in items:
        if r.group_id not in groups:
            grp_order.append(r.group_id)
        groups[r.group_id].append(r)

    def _grp_qty(grp_rows, otype):
        return next((r.quantity for r in grp_rows if r.output_type == otype), 0)

    total_plan   = sum((groups[g][0].production_plan or 0) for g in grp_order)
    total_fg     = sum(_grp_qty(groups[g], "finished_goods") for g in grp_order)
    total_actual = sum(sum(r.quantity for r in groups[g]) for g in grp_order)
    avg_q        = round(total_fg / total_actual * 100, 2) if total_actual else 0
    ws.row_dimensions[3].height = 46

    for idx, (lbl, val) in enumerate([
        ("Total Records", len(grp_order)), ("Plan (kg)", f"{total_plan:,}"),
        ("Actual (kg)", f"{total_actual:,}"), ("Good Product", f"{total_fg:,}"), ("Avg Quality", f"{avg_q}%")
    ]):
        cs = idx * 3 + 1
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs + 2)
        c = ws.cell(row=3, column=cs, value=f"{lbl}\n{val}")
        c.fill = PatternFill("solid", fgColor="DBEAFE")
        c.font = Font(bold=True, color="1D4ED8", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border("medium")

    ws.row_dimensions[4].height = 6

    COLS = [("No",5),("Tanggal",14),("Line",16),("Shift",12),("Kode Pakan",14),
            ("Plan (kg)",12),("FG (kg)",12),("DG (kg)",12),("WIP (kg)",12),
            ("Remix (kg)",12),("Reject (kg)",12),("Actual (kg)",12),
            ("Good Prod",12),("Quality %",12),("Remarks",24)]
    ws.row_dimensions[5].height = 30
    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=5, column=ci, value=h); _hdr(c, "3B82F6")
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, gid in enumerate(grp_order, 1):
        grp_rows = groups[gid]
        ref = grp_rows[0]
        fg  = _grp_qty(grp_rows, "finished_goods")
        dg  = _grp_qty(grp_rows, "downgraded_product")
        wip = _grp_qty(grp_rows, "wip")
        rmx = _grp_qty(grp_rows, "remix")
        rej = _grp_qty(grp_rows, "reject_product")
        act = fg + dg + wip + rmx + rej
        qp  = round(fg / act * 100, 2) if act else 0
        r = ri + 5; ws.row_dimensions[r].height = 18
        row_vals = [ri, ref.date.strftime("%d/%m/%Y") if ref.date else "",
            ref.line.name if ref.line else "", ref.shift.name if ref.shift else "",
            ref.feed_code.code if ref.feed_code else "",
            ref.production_plan or 0, fg, dg,
            wip, rmx, rej, act,
            fg, qp, ref.remarks or ""]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            _data(c, ri, "#,##0" if 6 <= ci <= 13 else None, center=ci <= 5)
            if ci == 14:
                c.number_format = '0.00"%"'
                if qp >= 98:
                    c.fill = PatternFill("solid", fgColor="D1FAE5"); c.font = Font(color="065F46", bold=True, size=9, name="Calibri")
                elif qp >= 95:
                    c.fill = PatternFill("solid", fgColor="FEF9C3"); c.font = Font(color="713F12", bold=True, size=9, name="Calibri")
                else:
                    c.fill = PatternFill("solid", fgColor="FEE2E2"); c.font = Font(color="7F1D1D", bold=True, size=9, name="Calibri")

    # Import Template sheet (dynamic columns berdasarkan active output types)
    ws2 = wb.create_sheet("Import Template")
    n_tmpl_cols = 6 + len(output_types) + 1  # fixed + output types + remarks
    _title_rows(ws2, "Import Template  |  Production Output", n_tmpl_cols, "059669")
    IH_FIXED = [("date\n(YYYY-MM-DD)",18),("line_name",18),("shift_name",18),
                ("feed_code\n(opsional)",16),("production_plan\n(kg)",16)]
    IH_OT    = [(f"{ot.code}\n(kg)", 16) for ot in output_types]
    IH_TRAIL = [("remarks\n(opsional)",24)]
    IH_ALL   = IH_FIXED + IH_OT + IH_TRAIL
    ws2.row_dimensions[3].height = 42
    for ci, (h, w) in enumerate(IH_ALL, 1):
        c = ws2.cell(row=3, column=ci, value=h); _hdr(c, "059669")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    sample_row = ["2025-01-15","LINE-01","Shift 1","P001",5000] + [100]*len(output_types) + ["Contoh data"]
    for ci, val in enumerate(sample_row, 1):
        c = ws2.cell(row=4, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor="F0FDF4"); c.font = Font(size=9, italic=True, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = _border()

    # Instructions sheet
    ws3 = wb.create_sheet("Petunjuk Import")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 5; ws3.column_dimensions["B"].width = 65
    t = ws3.cell(row=1, column=1, value="Petunjuk Penggunaan Import — Production Output")
    t.font = Font(bold=True, size=13, color="1D4ED8", name="Calibri"); ws3.row_dimensions[1].height = 28
    for i, text in enumerate([
        "Gunakan sheet 'Import Template' sebagai format data import.",
        "Kolom date harus format YYYY-MM-DD (contoh: 2025-01-15).",
        "line_name harus sesuai nama Line yang terdaftar di master.",
        "shift_name harus sesuai nama Shift yang terdaftar di master.",
        "feed_code boleh dikosongkan jika tidak ada kode pakan.",
        "Semua kolom qty (kg) harus berupa angka >= 0.",
        "Kolom remarks bersifat opsional.",
        "Baris header (baris ke-3) JANGAN dihapus.",
        "Hapus baris contoh (baris 4) sebelum import jika tidak dipakai.",
        "File harus dalam format .xlsx sebelum diupload.",
    ], start=3):
        ws3.cell(row=i, column=1, value=f"{i-2}.").font = Font(bold=True, color="1D4ED8", size=10, name="Calibri")
        c = ws3.cell(row=i, column=2, value=text)
        c.font = Font(size=10, name="Calibri"); ws3.row_dimensions[i].height = 20

    ws.freeze_panes = "A6"; ws.sheet_view.showGridLines = False
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"production_output_{plant.code}_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ════════════════════════════════════════════════════════
# PRODUCTION OUTPUTS – IMPORT
# ════════════════════════════════════════════════════════

@router.post("/production-outputs/import")
async def import_production_outputs(current_user: CurrentUser, plant: CurrentPlant, file: UploadFile = File(...)):
    """Import Production Outputs from Excel. Roles: administrator, plant_manager only."""
    _assert_role(current_user, IMPORT_ROLES, "Import Production Output")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "File harus berformat .xlsx")
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel tidak valid atau rusak")
    if "Import Template" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Import Template' tidak ditemukan. Gunakan template yang disediakan.")

    ws = wb["Import Template"]
    db = _db(plant)
    lines        = {l.name.strip().lower(): l for l in db.query(MasterLine).filter(MasterLine.is_active == True).all()}
    shifts       = {s.name.strip().lower(): s for s in db.query(MasterShift).filter(MasterShift.is_active == True).all()}
    fcodes       = {f.code.strip().lower(): f for f in db.query(MasterFeedCode).filter(MasterFeedCode.is_active == True).all()}
    output_types = _get_output_types(db)  # dynamic output types dari master

    # Kolom template: date(0), line(1), shift(2), fc(3), plan(4), ot[0..n-1](5..5+n-1), remarks(5+n)
    OT_START = 5

    created, errors = 0, []
    for rn, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(row): continue
        cells = list(row) + [None] * (OT_START + len(output_types) + 2)
        date_val, line_nm, shift_nm, fc_nm, plan = cells[0], cells[1], cells[2], cells[3], cells[4]
        remarks = cells[OT_START + len(output_types)]
        errs = []
        try:
            parsed_date = dt.fromisoformat(str(date_val).strip()) if isinstance(date_val, str) else dt.combine(date_val, dt.min.time())
        except:
            errs.append("format date salah"); parsed_date = None
        lo = lines.get((line_nm or "").strip().lower())
        so = shifts.get((shift_nm or "").strip().lower())
        if not lo: errs.append(f"line '{line_nm}' tidak ditemukan")
        if not so: errs.append(f"shift '{shift_nm}' tidak ditemukan")
        fco = fcodes.get(str(fc_nm or "").strip().lower()) if fc_nm else None
        if fc_nm and not fco: errs.append(f"kode pakan '{fc_nm}' tidak ditemukan")
        def si(v, lbl):
            try:
                x = int(float(v or 0))
                if x < 0: raise ValueError()
                return x
            except:
                errs.append(f"{lbl} harus angka >= 0"); return 0
        i_plan = si(plan, "production_plan")
        qty_map = {}
        for idx, ot in enumerate(output_types):
            qty_map[ot.code] = si(cells[OT_START + idx], ot.code)
        if errs: errors.append({"row": rn, "errors": errs}); continue
        gid = str(uuid.uuid4())
        for ot in output_types:
            db.add(ProductionOutput(
                group_id=gid,
                date=parsed_date, line_id=lo.id, shift_id=so.id,
                feed_code_id=fco.id if fco else None, production_plan=i_plan,
                output_type=ot.code,
                category=ot.category,
                quantity=qty_map.get(ot.code, 0),
                remarks=str(remarks).strip() if remarks else None,
                created_by_id=current_user.id,
            ))
        created += 1
    if created: db.commit()
    return {"imported": created, "errors": errors, "message": f"{created} baris berhasil diimport, {len(errors)} baris gagal."}


# ════════════════════════════════════════════════════════
# MACHINE LOSS INPUTS – CRUD
# ════════════════════════════════════════════════════════

@router.get("/machine-loss-inputs", response_model=list[MachineLossInputResponse])
def list_machine_loss_inputs(current_user: CurrentUser, plant: CurrentPlant,
    date_from: str | None = None, date_to: str | None = None,
    line_id: int | None = None, shift_id: int | None = None):
    db = _db(plant)
    q = db.query(MachineLossInput).filter(MachineLossInput.is_active == True)
    if date_from: q = q.filter(MachineLossInput.date >= dt.fromisoformat(date_from))
    if date_to:   q = q.filter(MachineLossInput.date <= dt.fromisoformat(date_to + "T23:59:59"))
    if line_id:   q = q.filter(MachineLossInput.line_id == line_id)
    if shift_id:  q = q.filter(MachineLossInput.shift_id == shift_id)
    return [_build_loss_response(i) for i in q.order_by(MachineLossInput.date.desc(), MachineLossInput.id.desc()).all()]


@router.post("/machine-loss-inputs", response_model=MachineLossInputResponse, status_code=201)
def create_machine_loss_input(payload: MachineLossInputCreate, current_user: CurrentUser, plant: CurrentPlant):
    db = _db(plant); _validate_loss_refs(db, payload.model_dump())
    item = MachineLossInput(date=payload.date, line_id=payload.line_id, shift_id=payload.shift_id,
        feed_code_id=payload.feed_code_id, loss_l1_id=payload.loss_l1_id,
        loss_l2_id=payload.loss_l2_id, loss_l3_id=payload.loss_l3_id,
        time_from=payload.time_from, time_to=payload.time_to,
        duration_minutes=payload.duration_minutes, remarks=payload.remarks, created_by_id=current_user.id)
    db.add(item); db.commit(); db.refresh(item); return _build_loss_response(item)


@router.put("/machine-loss-inputs/{item_id}", response_model=MachineLossInputResponse)
def update_machine_loss_input(item_id: int, payload: MachineLossInputUpdate, current_user: CurrentUser, plant: CurrentPlant):
    db = _db(plant)
    item = db.query(MachineLossInput).filter(MachineLossInput.id == item_id, MachineLossInput.is_active == True).first()
    if not item: raise HTTPException(404, "Machine loss input not found")
    data = payload.model_dump(exclude_none=True); _validate_loss_refs(db, data)
    for k, v in data.items(): setattr(item, k, v)
    item.updated_by_id = current_user.id; db.commit(); db.refresh(item); return _build_loss_response(item)


@router.delete("/machine-loss-inputs/{item_id}", status_code=204)
def delete_machine_loss_input(item_id: int, current_user: CurrentUser, plant: CurrentPlant):
    db = _db(plant)
    item = db.query(MachineLossInput).filter(MachineLossInput.id == item_id).first()
    if not item: raise HTTPException(404, "Machine loss input not found")
    item.is_active = False; item.updated_by_id = current_user.id; db.commit()


# ════════════════════════════════════════════════════════
# MACHINE LOSS INPUTS – EXPORT
# ════════════════════════════════════════════════════════

@router.get("/machine-loss-inputs/export")
def export_machine_loss_inputs(current_user: CurrentUser, plant: CurrentPlant,
    date_from: str | None = None, date_to: str | None = None,
    line_id: int | None = None, shift_id: int | None = None):
    """Export Machine Loss Inputs to Excel. Roles: administrator, plant_manager, operator."""
    _assert_role(current_user, EXPORT_ROLES, "Export Machine Loss")

    db = _db(plant)
    q = db.query(MachineLossInput).filter(MachineLossInput.is_active == True)
    if date_from: q = q.filter(MachineLossInput.date >= dt.fromisoformat(date_from))
    if date_to:   q = q.filter(MachineLossInput.date <= dt.fromisoformat(date_to + "T23:59:59"))
    if line_id:   q = q.filter(MachineLossInput.line_id == line_id)
    if shift_id:  q = q.filter(MachineLossInput.shift_id == shift_id)
    items = q.order_by(MachineLossInput.date.desc()).all()

    wb = Workbook(); ws = wb.active; ws.title = "Machine Loss Input"
    NCOLS = 13; _title_rows(ws, f"Machine Loss Input  |  {plant.name}", NCOLS, "7C3AED")

    total_dur = sum(i.duration_minutes for i in items)
    ws.row_dimensions[3].height = 46
    for idx, (lbl, val) in enumerate([
        ("Total Records", len(items)),
        ("Total Durasi (min)", f"{round(total_dur,1):,}"),
        ("Total Durasi (h:m)", f"{int(total_dur//60)}h {int(total_dur%60)}m")
    ]):
        cs = idx * 4 + 1
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=cs+3)
        c = ws.cell(row=3, column=cs, value=f"{lbl}\n{val}")
        c.fill = PatternFill("solid", fgColor="EDE9FE")
        c.font = Font(bold=True, color="6D28D9", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border("medium")

    ws.row_dimensions[4].height = 6
    COLS = [("No",4),("Tanggal",14),("Line",16),("Shift",12),("Kode Pakan",14),
            ("Loss L1",22),("Loss L2",22),("Loss L3",22),("Dari",8),("Sampai",8),
            ("Durasi (min)",14),("Durasi (h:m)",14),("Remarks",24)]
    ws.row_dimensions[5].height = 30
    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=5, column=ci, value=h); _hdr(c, "8B5CF6")
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, item in enumerate(items, 1):
        r = ri + 5; ws.row_dimensions[r].height = 18
        dur = item.duration_minutes
        row_vals = [ri, item.date.strftime("%d/%m/%Y") if item.date else "",
            item.line.name if item.line else "", item.shift.name if item.shift else "",
            item.feed_code.code if item.feed_code else "",
            item.loss_l1.name if item.loss_l1 else "", item.loss_l2.name if item.loss_l2 else "",
            item.loss_l3.name if item.loss_l3 else "",
            item.time_from or "", item.time_to or "", dur,
            f"{int(dur//60)}h {int(dur%60):02d}m", item.remarks or ""]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            _data(c, ri, "#,##0.0" if ci == 11 else None, center=ci in {1,9,10,11,12})
            if ci == 11:
                if dur <= 30:
                    c.fill = PatternFill("solid", fgColor="D1FAE5"); c.font = Font(color="065F46", bold=True, size=9, name="Calibri")
                elif dur <= 120:
                    c.fill = PatternFill("solid", fgColor="FEF9C3"); c.font = Font(color="713F12", bold=True, size=9, name="Calibri")
                else:
                    c.fill = PatternFill("solid", fgColor="FEE2E2"); c.font = Font(color="7F1D1D", bold=True, size=9, name="Calibri")

    # Import Template sheet
    ws2 = wb.create_sheet("Import Template")
    _title_rows(ws2, "Import Template  |  Machine Loss Input", 11, "059669")
    IH = [("date\n(YYYY-MM-DD)",18),("line_name",18),("shift_name",18),
          ("feed_code\n(opsional)",16),("loss_l1_name\n(opsional)",22),
          ("loss_l2_name\n(opsional)",22),("loss_l3_name\n(opsional)",22),
          ("time_from\n(HH:MM)",16),("time_to\n(HH:MM)",16),
          ("duration_minutes\n(wajib, >0)",20),("remarks\n(opsional)",22)]
    ws2.row_dimensions[3].height = 42
    for ci, (h, w) in enumerate(IH, 1):
        c = ws2.cell(row=3, column=ci, value=h); _hdr(c, "059669")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    for ci, val in enumerate(["2025-01-15","LINE-01","Shift 1","P001","Planned Stop","Changeover","Cleaning","08:00","09:30",90.0,"Ganti produk"], 1):
        c = ws2.cell(row=4, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor="F0FDF4"); c.font = Font(size=9, italic=True, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = _border()

    ws.freeze_panes = "A6"; ws.sheet_view.showGridLines = False
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"machine_loss_{plant.code}_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ════════════════════════════════════════════════════════
# MACHINE LOSS INPUTS – IMPORT
# ════════════════════════════════════════════════════════

@router.post("/machine-loss-inputs/import")
async def import_machine_loss_inputs(current_user: CurrentUser, plant: CurrentPlant, file: UploadFile = File(...)):
    """Import Machine Loss Inputs from Excel. Roles: administrator, plant_manager only."""
    _assert_role(current_user, IMPORT_ROLES, "Import Machine Loss")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "File harus berformat .xlsx")
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except:
        raise HTTPException(400, "File Excel tidak valid atau rusak")
    if "Import Template" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Import Template' tidak ditemukan. Gunakan template yang disediakan.")

    ws = wb["Import Template"]
    db = _db(plant)
    lines  = {l.name.strip().lower(): l for l in db.query(MasterLine).filter(MasterLine.is_active == True).all()}
    shifts = {s.name.strip().lower(): s for s in db.query(MasterShift).filter(MasterShift.is_active == True).all()}
    fcodes = {f.code.strip().lower(): f for f in db.query(MasterFeedCode).filter(MasterFeedCode.is_active == True).all()}
    losses_l1 = {l.name.strip().lower(): l for l in db.query(MachineLossLvl1).all()}
    losses_l2 = {l.name.strip().lower(): l for l in db.query(MachineLossLvl2).all()}
    losses_l3 = {l.name.strip().lower(): l for l in db.query(MachineLossLvl3).all()}

    created, errors = 0, []
    for rn, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(row): continue
        cells = list(row) + [None]*11
        date_val,line_nm,shift_nm,fc_nm,l1_nm,l2_nm,l3_nm,tf,tt,dur_raw,remarks = cells[:11]
        errs = []
        try:
            parsed_date = dt.fromisoformat(str(date_val).strip()) if isinstance(date_val, str) else dt.combine(date_val, dt.min.time())
        except:
            errs.append("format date salah"); parsed_date = None
        lo = lines.get((line_nm or "").strip().lower())
        so = shifts.get((shift_nm or "").strip().lower())
        if not lo: errs.append(f"line '{line_nm}' tidak ditemukan")
        if not so: errs.append(f"shift '{shift_nm}' tidak ditemukan")
        fco = fcodes.get(str(fc_nm or "").strip().lower()) if fc_nm else None
        l1o = losses_l1.get(str(l1_nm or "").strip().lower()) if l1_nm else None
        l2o = losses_l2.get(str(l2_nm or "").strip().lower()) if l2_nm else None
        l3o = losses_l3.get(str(l3_nm or "").strip().lower()) if l3_nm else None
        if fc_nm and not fco: errs.append(f"kode pakan '{fc_nm}' tidak ditemukan")
        if l1_nm and not l1o: errs.append(f"loss L1 '{l1_nm}' tidak ditemukan")
        if l2_nm and not l2o: errs.append(f"loss L2 '{l2_nm}' tidak ditemukan")
        if l3_nm and not l3o: errs.append(f"loss L3 '{l3_nm}' tidak ditemukan")
        try:
            dur = float(dur_raw or 0)
            if dur <= 0: raise ValueError()
        except:
            errs.append("duration_minutes harus angka > 0"); dur = None
        if errs: errors.append({"row": rn, "errors": errs}); continue
        db.add(MachineLossInput(date=parsed_date, line_id=lo.id, shift_id=so.id,
            feed_code_id=fco.id if fco else None,
            loss_l1_id=l1o.machine_losses_lvl_1_id if l1o else None,
            loss_l2_id=l2o.machine_losses_lvl_2_id if l2o else None,
            loss_l3_id=l3o.machine_losses_lvl_3_id if l3o else None,
            time_from=str(tf).strip() if tf else None, time_to=str(tt).strip() if tt else None,
            duration_minutes=dur, remarks=str(remarks).strip() if remarks else None,
            created_by_id=current_user.id))
        created += 1
    if created: db.commit()
    return {"imported": created, "errors": errors, "message": f"{created} baris berhasil diimport, {len(errors)} baris gagal."}