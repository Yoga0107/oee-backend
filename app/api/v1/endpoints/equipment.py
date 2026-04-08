
import io
from datetime import datetime as dt

from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.deps import CurrentUser, CurrentPlant
from app.db.database import get_plant_db
from app.models.plant_schema import EquipmentTree
from app.schemas.equipment import EquipmentTreeCreate, EquipmentTreeUpdate, EquipmentTreeResponse

router = APIRouter(prefix="/equipment", tags=["Equipment Tree"])

WRITE_ROLES  = {"administrator", "plant_manager"}
VERIFY_ROLES = {"administrator", "plant_manager"}


def _db(plant: CurrentPlant) -> Session:
    return next(get_plant_db(plant.schema_name))


def _assert_role(user, roles: set, action: str = "action"):
    role_name = user.role.name if user.role else ""
    if not user.is_superuser and role_name not in roles:
        raise HTTPException(403, f"Akses ditolak: {action} hanya untuk {', '.join(sorted(roles))}")


def _border(style="thin"):
    s = Side(style=style, color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg="1F4E78"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("medium")


def _apply_filters(q, sistem, sub_sistem, unit_mesin, bagian_mesin, is_verified, search):
    """Apply shared filter logic — reused by list and paginated endpoints."""
    if sistem:       q = q.filter(EquipmentTree.sistem.ilike(f"%{sistem}%"))
    if sub_sistem:   q = q.filter(EquipmentTree.sub_sistem.ilike(f"%{sub_sistem}%"))
    if unit_mesin:   q = q.filter(EquipmentTree.unit_mesin.ilike(f"%{unit_mesin}%"))
    if bagian_mesin: q = q.filter(EquipmentTree.bagian_mesin.ilike(f"%{bagian_mesin}%"))
    if is_verified is not None:
        q = q.filter(EquipmentTree.is_verified == is_verified)
    if search:
        like = f"%{search}%"
        q = q.filter(
            EquipmentTree.sistem.ilike(like)       |
            EquipmentTree.sub_sistem.ilike(like)   |
            EquipmentTree.unit_mesin.ilike(like)   |
            EquipmentTree.bagian_mesin.ilike(like) |
            EquipmentTree.spare_part.ilike(like)   |
            EquipmentTree.spesifikasi.ilike(like)  |
            EquipmentTree.sku.ilike(like)
        )
    return q


_ORDER = [
    EquipmentTree.sistem,
    EquipmentTree.sub_sistem,
    EquipmentTree.unit_mesin,
    EquipmentTree.bagian_mesin,
    EquipmentTree.spare_part,
]


# ─── Paginated list (utama — untuk table view) ────────────────────────────────

@router.get("/paginated")
def list_equipment_paginated(
    response:     Response,
    current_user: CurrentUser,
    plant:        CurrentPlant,
    sistem:       str | None  = Query(None),
    sub_sistem:   str | None  = Query(None),
    unit_mesin:   str | None  = Query(None),
    bagian_mesin: str | None  = Query(None),
    is_verified:  bool | None = Query(None),
    search:       str | None  = Query(None),
    page:         int         = Query(1, ge=1),
    page_size:    int         = Query(25, ge=1, le=100),
):
    """
    Server-side pagination. Returns:
      - data: list of items for current page
      - total: total matching records (for pagination UI)
      - page, page_size, total_pages

    Hanya fetch data yang dibutuhkan — tidak load semua rows ke memori.
    """
    db = _db(plant)
    q  = db.query(EquipmentTree).filter(EquipmentTree.is_active == True)
    q  = _apply_filters(q, sistem, sub_sistem, unit_mesin, bagian_mesin, is_verified, search)

    # Count query (ringan — no SELECT *)
    total = q.with_entities(func.count(EquipmentTree.id)).scalar()

    total_pages = max(1, -(-total // page_size))   # ceil division
    skip        = (page - 1) * page_size

    items = q.order_by(*_ORDER).offset(skip).limit(page_size).all()

    return {
        "data":        [EquipmentTreeResponse.model_validate(i) for i in items],
        "total":       total,
        "page":        page,
        "page_size":   page_size,
        "total_pages": total_pages,
    }


# ─── Full list (untuk tree view — load semua, tapi hanya field minimal) ───────

@router.get("", response_model=list[EquipmentTreeResponse])
def list_equipment(
    current_user: CurrentUser,
    plant:        CurrentPlant,
    sistem:       str | None  = Query(None),
    sub_sistem:   str | None  = Query(None),
    unit_mesin:   str | None  = Query(None),
    bagian_mesin: str | None  = Query(None),
    is_verified:  bool | None = Query(None),
    search:       str | None  = Query(None),
    skip:         int         = Query(0, ge=0),
    limit:        int         = Query(500, ge=1, le=2000),
):
    """Full list — dipakai untuk tree view. Limit default 500."""
    db = _db(plant)
    q  = db.query(EquipmentTree).filter(EquipmentTree.is_active == True)
    q  = _apply_filters(q, sistem, sub_sistem, unit_mesin, bagian_mesin, is_verified, search)
    return q.order_by(*_ORDER).offset(skip).limit(limit).all()


# ─── Stats ────────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(current_user: CurrentUser, plant: CurrentPlant):
    db = _db(plant)
    total     = db.query(func.count(EquipmentTree.id)).filter(EquipmentTree.is_active == True).scalar()
    verified  = db.query(func.count(EquipmentTree.id)).filter(EquipmentTree.is_active == True, EquipmentTree.is_verified == True).scalar()
    by_sistem = (
        db.query(EquipmentTree.sistem, func.count(EquipmentTree.id))
        .filter(EquipmentTree.is_active == True)
        .group_by(EquipmentTree.sistem)
        .order_by(EquipmentTree.sistem)
        .all()
    )
    return {
        "total":      total,
        "verified":   verified,
        "unverified": total - verified,
        "by_sistem":  [{"sistem": s, "count": c} for s, c in by_sistem],
    }
    
    # ─── Tambahkan endpoint ini ke: oee-backend-master/app/api/v1/endpoints/equipment.py
# Letakkan setelah endpoint /stats yang sudah ada
# Pastikan import datetime sudah ada (sudah ada di file asli)
# Tambahkan import tambahan di bagian atas file jika belum ada:
#   from datetime import timedelta
#   from collections import defaultdict


# ─── Trend Analysis ───────────────────────────────────────────────────────────

@router.get("/trend-analysis")
def get_trend_analysis(
    current_user: CurrentUser,
    plant:        CurrentPlant,
    period:       str = Query("monthly", regex="^(daily|weekly|monthly|quarterly|yearly)$"),
    start_date:   str | None = Query(None, description="Format: YYYY-MM-DD"),
    end_date:     str | None = Query(None, description="Format: YYYY-MM-DD"),
    sistem:       str | None = Query(None),
    group_by:     str = Query("level", regex="^(level|sistem|bu|verified)$"),
):
    """
    Trend Analysis untuk pengumpulan data Equipment.
    
    Menampilkan:
    - Jumlah data yang ditambahkan per periode waktu
    - Perbandingan antar periode (apakah ada peningkatan/penurunan)
    - Peak period (periode dengan penambahan terbanyak)
    - Breakdown berdasarkan group_by (level, sistem, bu, atau status verifikasi)
    - Growth rate per periode
    
    Params:
    - period: daily | weekly | monthly | quarterly | yearly
    - start_date: tanggal mulai filter (YYYY-MM-DD)
    - end_date: tanggal akhir filter (YYYY-MM-DD)
    - sistem: filter berdasarkan sistem tertentu
    - group_by: level | sistem | bu | verified
    """
    from datetime import timedelta
    from collections import defaultdict
    import calendar

    db = _db(plant)

    # Base query — hanya ambil data aktif
    q = db.query(EquipmentTree).filter(EquipmentTree.is_active == True)

    # Filter sistem jika ada
    if sistem:
        q = q.filter(EquipmentTree.sistem.ilike(f"%{sistem}%"))

    # Parse date range
    try:
        dt_start = dt.strptime(start_date, "%Y-%m-%d") if start_date else None
        dt_end   = dt.strptime(end_date,   "%Y-%m-%d").replace(hour=23, minute=59, second=59) if end_date else None
    except ValueError:
        raise HTTPException(400, "Format tanggal tidak valid. Gunakan YYYY-MM-DD")

    if dt_start:
        q = q.filter(EquipmentTree.created_at >= dt_start)
    if dt_end:
        q = q.filter(EquipmentTree.created_at <= dt_end)

    items = q.order_by(EquipmentTree.created_at).all()

    if not items:
        return {
            "period":       period,
            "group_by":     group_by,
            "total_records": 0,
            "timeline":     [],
            "summary":      {},
            "peak_period":  None,
            "growth_trend": [],
            "breakdown":    [],
        }

    # ── Helper: format period key ─────────────────────────────────────────────
    def get_period_key(d: dt) -> str:
        if period == "daily":
            return d.strftime("%Y-%m-%d")
        elif period == "weekly":
            # ISO week: YYYY-WNN
            year, week, _ = d.isocalendar()
            return f"{year}-W{week:02d}"
        elif period == "monthly":
            return d.strftime("%Y-%m")
        elif period == "quarterly":
            q_num = (d.month - 1) // 3 + 1
            return f"{d.year}-Q{q_num}"
        else:  # yearly
            return str(d.year)

    def get_period_label(key: str) -> str:
        """Ubah period key menjadi label yang lebih readable."""
        if period == "daily":
            try:
                d = dt.strptime(key, "%Y-%m-%d")
                return d.strftime("%d %b %Y")
            except:
                return key
        elif period == "weekly":
            # "2024-W03" → "Minggu 3, 2024"
            parts = key.split("-W")
            if len(parts) == 2:
                return f"Minggu {parts[1]}, {parts[0]}"
            return key
        elif period == "monthly":
            try:
                d = dt.strptime(key, "%Y-%m")
                return d.strftime("%B %Y")
            except:
                return key
        elif period == "quarterly":
            parts = key.split("-")
            if len(parts) == 2:
                return f"{parts[1]} {parts[0]}"
            return key
        else:  # yearly
            return key

    # ── Grouping helper untuk breakdown ──────────────────────────────────────
    def get_group_key(item: EquipmentTree) -> str:
        if group_by == "level":
            # Tentukan level paling dalam yang terisi
            if item.spare_part:    return "spare_part"
            if item.bagian_mesin:  return "bagian_mesin"
            if item.unit_mesin:    return "unit_mesin"
            if item.sub_sistem:    return "sub_sistem"
            return "sistem"
        elif group_by == "sistem":
            return item.sistem or "—"
        elif group_by == "bu":
            return item.bu or "Tanpa BU"
        elif group_by == "verified":
            return "Verified" if item.is_verified else "Pending"
        return "—"

    # ── Build timeline data ───────────────────────────────────────────────────
    timeline_counts: dict[str, int] = defaultdict(int)
    timeline_breakdown: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    # Kumpulkan semua verified_at juga untuk trend verifikasi
    verify_timeline: dict[str, int] = defaultdict(int)

    for item in items:
        pk = get_period_key(item.created_at)
        timeline_counts[pk] += 1
        gk = get_group_key(item)
        timeline_breakdown[pk][gk] += 1

        # Track verifikasi berdasarkan verified_at
        if item.verified_at:
            vpk = get_period_key(item.verified_at)
            verify_timeline[vpk] += 1

    # Urutkan berdasarkan key (kronologis)
    sorted_periods = sorted(timeline_counts.keys())

    # ── Hitung growth rate ────────────────────────────────────────────────────
    growth_trend = []
    prev_count   = None
    for pk in sorted_periods:
        count = timeline_counts[pk]
        if prev_count is not None and prev_count > 0:
            growth = round(((count - prev_count) / prev_count) * 100, 1)
        elif prev_count == 0 and count > 0:
            growth = 100.0
        else:
            growth = 0.0
        growth_trend.append({
            "period":      pk,
            "label":       get_period_label(pk),
            "count":       count,
            "growth_rate": growth,
            "trend":       "up" if growth > 0 else ("down" if growth < 0 else "stable"),
        })
        prev_count = count

    # ── Peak period ───────────────────────────────────────────────────────────
    peak_period = None
    if timeline_counts:
        peak_key   = max(timeline_counts, key=lambda k: timeline_counts[k])
        peak_count = timeline_counts[peak_key]
        peak_period = {
            "period": peak_key,
            "label":  get_period_label(peak_key),
            "count":  peak_count,
        }

    # ── Build breakdown list (semua grup di semua periode) ────────────────────
    all_groups = sorted(set(
        gk for period_data in timeline_breakdown.values() for gk in period_data
    ))

    # Label mapping untuk level
    level_labels = {
        "sistem":       "Sistem",
        "sub_sistem":   "Sub Sistem",
        "unit_mesin":   "Unit Mesin",
        "bagian_mesin": "Bagian Mesin",
        "spare_part":   "Spare Part",
    }

    breakdown_series = []
    for gk in all_groups:
        series_data = []
        for pk in sorted_periods:
            series_data.append({
                "period": pk,
                "label":  get_period_label(pk),
                "count":  timeline_breakdown[pk].get(gk, 0),
            })
        label = level_labels.get(gk, gk) if group_by == "level" else gk
        total_in_group = sum(d["count"] for d in series_data)
        breakdown_series.append({
            "group":  gk,
            "label":  label,
            "total":  total_in_group,
            "series": series_data,
        })

    # Urutkan breakdown berdasarkan total terbanyak
    breakdown_series.sort(key=lambda x: x["total"], reverse=True)

    # ── Summary statistics ────────────────────────────────────────────────────
    counts         = [timeline_counts[pk] for pk in sorted_periods]
    total_records  = sum(counts)
    avg_per_period = round(total_records / len(counts), 1) if counts else 0
    max_count      = max(counts) if counts else 0
    min_count      = min(counts) if counts else 0

    # Overall trend: bandingkan paruh pertama vs paruh kedua
    if len(counts) >= 2:
        mid       = len(counts) // 2
        first_avg = sum(counts[:mid]) / mid if mid > 0 else 0
        last_avg  = sum(counts[mid:]) / (len(counts) - mid) if (len(counts) - mid) > 0 else 0
        if last_avg > first_avg * 1.05:
            overall_trend = "increasing"
        elif last_avg < first_avg * 0.95:
            overall_trend = "decreasing"
        else:
            overall_trend = "stable"
    else:
        overall_trend = "stable"

    # ── Full timeline (untuk chart utama) ─────────────────────────────────────
    timeline = []
    cumulative = 0
    for pk in sorted_periods:
        count = timeline_counts[pk]
        cumulative += count
        timeline.append({
            "period":           pk,
            "label":            get_period_label(pk),
            "count":            count,
            "cumulative":       cumulative,
            "verified_count":   verify_timeline.get(pk, 0),
        })

    return {
        "period":        period,
        "group_by":      group_by,
        "total_records": total_records,
        "date_range": {
            "start": items[0].created_at.isoformat() if items else None,
            "end":   items[-1].created_at.isoformat() if items else None,
        },
        "timeline":      timeline,
        "summary": {
            "total_periods":   len(sorted_periods),
            "avg_per_period":  avg_per_period,
            "max_per_period":  max_count,
            "min_per_period":  min_count,
            "overall_trend":   overall_trend,
        },
        "peak_period":  peak_period,
        "growth_trend": growth_trend,
        "breakdown":    breakdown_series,
    }


# ─── Create ───────────────────────────────────────────────────────────────────

@router.post("", response_model=EquipmentTreeResponse, status_code=201)
def create_equipment(payload: EquipmentTreeCreate, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, WRITE_ROLES, "Create Equipment")
    db   = _db(plant)
    item = EquipmentTree(**payload.model_dump(), created_by_id=current_user.id)
    db.add(item); db.commit(); db.refresh(item)
    return item


# ─── Update ───────────────────────────────────────────────────────────────────

@router.put("/{item_id}", response_model=EquipmentTreeResponse)
def update_equipment(item_id: int, payload: EquipmentTreeUpdate, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, WRITE_ROLES, "Update Equipment")
    db   = _db(plant)
    item = db.query(EquipmentTree).filter(EquipmentTree.id == item_id, EquipmentTree.is_active == True).first()
    if not item:
        raise HTTPException(404, "Data tidak ditemukan")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    item.updated_by_id = current_user.id
    if any(k in payload.model_dump(exclude_unset=True) for k in
           ("sistem","sub_sistem","unit_mesin","bagian_mesin","spare_part","spesifikasi","sku","bu")):
        item.is_verified    = False
        item.verified_by_id = None
        item.verified_at    = None
    db.commit(); db.refresh(item)
    return item


# ─── Delete (soft) ────────────────────────────────────────────────────────────

@router.delete("/{item_id}", status_code=204)
def delete_equipment(item_id: int, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, WRITE_ROLES, "Delete Equipment")
    db   = _db(plant)
    item = db.query(EquipmentTree).filter(EquipmentTree.id == item_id, EquipmentTree.is_active == True).first()
    if not item:
        raise HTTPException(404, "Data tidak ditemukan")
    item.is_active     = False
    item.updated_by_id = current_user.id
    db.commit()


# ─── Verify / Unverify ────────────────────────────────────────────────────────

@router.post("/{item_id}/verify", response_model=EquipmentTreeResponse)
def verify_equipment(item_id: int, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, VERIFY_ROLES, "Verify Equipment")
    db   = _db(plant)
    item = db.query(EquipmentTree).filter(EquipmentTree.id == item_id, EquipmentTree.is_active == True).first()
    if not item: raise HTTPException(404, "Data tidak ditemukan")
    item.is_verified    = True
    item.verified_by_id = current_user.id
    item.verified_at    = dt.utcnow()
    item.updated_by_id  = current_user.id
    db.commit(); db.refresh(item)
    return item


@router.post("/{item_id}/unverify", response_model=EquipmentTreeResponse)
def unverify_equipment(item_id: int, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, VERIFY_ROLES, "Unverify Equipment")
    db   = _db(plant)
    item = db.query(EquipmentTree).filter(EquipmentTree.id == item_id, EquipmentTree.is_active == True).first()
    if not item: raise HTTPException(404, "Data tidak ditemukan")
    item.is_verified    = False
    item.verified_by_id = None
    item.verified_at    = None
    item.updated_by_id  = current_user.id
    db.commit(); db.refresh(item)
    return item


@router.post("/verify-bulk")
def verify_bulk(payload: dict, current_user: CurrentUser, plant: CurrentPlant):
    _assert_role(current_user, VERIFY_ROLES, "Bulk Verify")
    ids    = payload.get("ids", [])
    action = payload.get("action", "verify")
    if not ids: raise HTTPException(400, "ids tidak boleh kosong")
    db    = _db(plant)
    items = db.query(EquipmentTree).filter(EquipmentTree.id.in_(ids), EquipmentTree.is_active == True).all()
    now   = dt.utcnow()
    for item in items:
        if action == "verify":
            item.is_verified = True; item.verified_by_id = current_user.id; item.verified_at = now
        else:
            item.is_verified = False; item.verified_by_id = None; item.verified_at = None
        item.updated_by_id = current_user.id
    db.commit()
    return {"updated": len(items), "action": action}


# ─── Export Excel ─────────────────────────────────────────────────────────────

@router.get("/export")
def export_equipment(current_user: CurrentUser, plant: CurrentPlant, is_verified: bool | None = Query(None)):
    db = _db(plant)
    q  = db.query(EquipmentTree).filter(EquipmentTree.is_active == True)
    if is_verified is not None:
        q = q.filter(EquipmentTree.is_verified == is_verified)
    items = q.order_by(*_ORDER).all()

    wb = Workbook(); ws = wb.active; ws.title = "Equipment Tree"
    COLS = [
        ("Sistem", 20), ("Sub Sistem", 22), ("Unit Mesin", 25),
        ("Bagian Mesin", 22), ("Spare Part", 25), ("Spesifikasi", 35),
        ("SKU", 18), ("BU", 10), ("Verified", 10), ("Remarks", 25),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=h); _hdr(c)
        ws.column_dimensions[get_column_letter(ci)].width = w

    VF = PatternFill("solid", fgColor="D1FAE5")
    PF = PatternFill("solid", fgColor="FEF9C3")
    for ri, item in enumerate(items, 2):
        vals = [item.sistem, item.sub_sistem, item.unit_mesin, item.bagian_mesin,
                item.spare_part, item.spesifikasi, item.sku, item.bu,
                "✓ Verified" if item.is_verified else "Pending", item.remarks]
        fill = VF if item.is_verified else PF
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val or "")
            c.border = _border(); c.alignment = Alignment(vertical="center", wrap_text=(ci == 6))
            if ci == 9:
                c.fill = fill
                c.font = Font(color="065F46" if item.is_verified else "713F12", bold=True, size=9, name="Calibri")

    ws2 = wb.create_sheet("Import Template")
    IH  = [("sistem*",20),("sub_sistem",22),("unit_mesin",25),("bagian_mesin",22),
           ("spare_part",25),("spesifikasi",35),("sku",18),("bu",10),("remarks",25)]
    for ci, (h, w) in enumerate(IH, 1):
        c = ws2.cell(row=1, column=ci, value=h); _hdr(c, "059669")
        ws2.column_dimensions[get_column_letter(ci)].width = w
    example = ["INTAKE","FINE INTAKE","Blower Fan - BF 111","Motor","Motor",
               "Teco AEEBKBO20020FMB 15 kW 3000 rpm","","FF",""]
    for ci, val in enumerate(example, 1):
        c = ws2.cell(row=2, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor="F0FDF4")
        c.font = Font(size=9, italic=True, name="Calibri"); c.border = _border()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"equipment_tree_{plant.name}_{dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ─── Import Excel ─────────────────────────────────────────────────────────────

@router.post("/import")
async def import_equipment(current_user: CurrentUser, plant: CurrentPlant, file: UploadFile = File(...)):
    _assert_role(current_user, WRITE_ROLES, "Import Equipment")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "File harus .xlsx")
    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "File Excel tidak valid")
    if "Import Template" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Import Template' tidak ditemukan.")
    ws = wb["Import Template"]
    db = _db(plant)
    created, errors = 0, []
    for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        cells = list(row) + [None] * 9
        sistem, sub_sistem, unit_mesin, bagian_mesin, spare_part, \
            spesifikasi, sku, bu, remarks = [
            str(c).strip() if c is not None else None for c in cells[:9]]
        if not sistem:
            errors.append({"row": rn, "errors": ["kolom 'sistem' wajib diisi"]}); continue
        db.add(EquipmentTree(
            sistem=sistem, sub_sistem=sub_sistem, unit_mesin=unit_mesin,
            bagian_mesin=bagian_mesin, spare_part=spare_part,
            spesifikasi=spesifikasi, sku=sku, bu=bu, remarks=remarks,
            created_by_id=current_user.id))
        created += 1
    if created: db.commit()
    return {"imported": created, "errors": errors,
            "message": f"{created} baris berhasil diimport, {len(errors)} baris gagal."}
